#!/usr/bin/env python 
# encoding: utf-8
import socket
import wmi
import pathlib
import time
import os
import winreg
from openpyxl import Workbook #不需要excel文件存在 可以生成一个
from openpyxl import load_workbook #需要传入一个excel文件不能传入


#计算机名
def printHostname():
    hostname = socket.gethostname()
    # print(hostname)
    return hostname

#CPU
def printCPU():
    CPU = []
    for cpu in c.Win32_Processor():
        Cpu_dict = {}
        Cpu_dict['CPU型号'] = cpu.Name
        CPU.append(Cpu_dict)
    # print(CPU)
    return CPU

#BIOS
def printBIOS():
    BIOS = []
    for bios in c.Win32_BIOS():
        BIOS_dict = {}
        BIOS_dict['服务编码'] = bios.SerialNumber
        BIOS.append(BIOS_dict)
    # print(BIOS)
    return BIOS

#内存
def printPhysicalMemory():
    memorys = []

    for mem in c.Win32_PhysicalMemory():
        mem_dict = {}
        mem_dict['内存插槽'] = mem.DeviceLocator
        mem_dict['容量'] = int(int(mem.Capacity)/1024/1024)
        mem_dict['速度'] = mem.Speed
        mem_dict['厂商'] = mem.Manufacturer
        mem_dict['零件号'] = mem.PartNumber
        memorys.append(mem_dict)
    # for m in memorys:
    #     print(m)
    # print(memorys)
    return memorys

#硬盘
def printDisk():
    disks = []
    for disk in c.Win32_DiskDrive():
        tmpmsg = {}
        tmpmsg['硬盘名称'] = disk.Caption
        tmpmsg['服务编码'] = disk.SerialNumber.strip()
        tmpmsg['容量'] = int(int(disk.Size)/1024/1024/1024)
        disks.append(tmpmsg)
    # for d in disks:
    #     print(d)
    # print(disks)
    return disks

#显卡
def printDisplay_Card():
    Display_Card = []
    for dp in c.Win32_VideoController():
        Display_Card_dict = {}
        Display_Card_dict['显卡名称'] = dp.Caption
        Display_Card.append(Display_Card_dict)
    # print(Display_Card)
    return Display_Card

#网卡mac地址
def printMacAddress():
    macs = []
    for n in  c.Win32_NetworkAdapter():
        mactmp = n.MACAddress
        if mactmp and len(mactmp.strip()) > 5:
            tmpmsg = {}
            tmpmsg['网卡名称'] = n.Name
            tmpmsg['MAC地址'] = n.MACAddress
            macs.append(tmpmsg)
    # print(macs)
    return macs

#操作系统
def printOS():
    OS = []
    for o in c.Win32_OperatingSystem():
        OS_dict = {}
        OS_dict['操作系统'] = o.Caption
        OS.append(OS_dict)
    # print(BIOS)
    return OS

#excel生成
def excel_new(first_line,excel_name):

    # 获取工作簿
    wb = Workbook()
    # 激活当前的sheet
    ws = wb.active
    ws.append(first_line)
    ws1 = wb.create_sheet('SoftwareList')
    wb.save(excel_name)


#ecel表格处理
def excel_handle(Infolist,SoftList,path_0,excel_name):
    try:

        # 将信息写入excel表格
        wb = load_workbook(path_0 + '\\' + excel_name)
        # 获取所有sheet
        sheets = wb.sheetnames
        #获取第一个sheet
        ws = wb[sheets[0]]
        #添加行
        ws.append(Infolist)
        #获取第二个sheet
        ws2 = wb[sheets[1]]
        #插入一列
        ws2.insert_cols(idx=1,amount=2)
        #在第2行第一列 添加value为 SoftList[i] 的参数
        for i in range(len(SoftList)):
            ws2.cell(i + 1, 1).value = SoftList[i]

        # 将此属性设置为False（默认），以保存为文档：
        wb.template = False
        # 您可以指定属性template = True，以将工作簿另存为模板：
        # wb.template = True
        #保存
        wb.save(filename= path_0 + '\\' + excel_name)
    except Exception as e:
        print("Exception", e)
        input('写入excel表格报错了！！！')


#列出软件列表
def SoftwareList():
    try:

        # 格式化pathlib返回的当前路径
        path_0 = str(pathlib.Path.cwd())

        # 使用主机名命名软件安装列表
        hostname = socket.gethostname()

        # file = open(path_0 + '\\' + '%s' % hostname + '.txt', 'w+')

        # 需要遍历的两个注册表
        sub_key = [r'SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall',
                   r'SOFTWARE\Wow6432Node\Microsoft\Windows\CurrentVersion\Uninstall']

        software_name = []

        for i in sub_key:
            key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, i, 0, winreg.KEY_ALL_ACCESS)
            for j in range(0, winreg.QueryInfoKey(key)[0] - 1):
                try:
                    key_name = winreg.EnumKey(key, j)
                    key_path = i + '\\' + key_name
                    each_key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, key_path, 0, winreg.KEY_ALL_ACCESS)
                    DisplayName, REG_SZ = winreg.QueryValueEx(each_key, 'DisplayName')
                    DisplayName = DisplayName.encode('utf-8')
                    software_name.append(DisplayName)
                except WindowsError:
                    pass

        # 去重排序
        software_name = list(set(software_name))
        software_name = sorted(software_name)
        SoftList = []
        for result in software_name:
            # print(result.decode("utf-8"))
            # file.write(str(result.decode("utf-8")) + '\n')
            SoftList.append(str(result.decode("utf-8")))
        # file.close()
        return SoftList
    except Exception as e:
        print("Exception", e)
        input('获取软件列表报错了！！')


def Info_list():
    #使用人
    use_name = input("请输入用户姓名,以回车键结束。")
    # print(use_name)
    #计算机名
    hostname = printHostname()
    # print(hostname)
    #SN
    tmpSN = printBIOS()
    for i in tmpSN:
        SN = i['服务编码']
    # print(SN)
    #操作系统
    tmpOS = printOS()
    for i in tmpOS:
        OS = i['操作系统']
    # print(OS)
    #CPU
    tmpCPU = printCPU()
    for i in tmpCPU:
        CPU = i['CPU型号']
    # print(CPU)
    #内存
    Mem = str(printPhysicalMemory())
    # print(printPhysicalMemory())
    #硬盘
    Hdisk = str(printDisk())
    # print(printDisk())
    #显卡
    Display_Card = str(printDisplay_Card())
    # print(printDisplay_Card())
    #网卡
    MacAddress = str(printMacAddress())
    # print(printMacAddress())
    #导入时间
    Import_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    Infolist = [use_name,hostname,SN,OS,CPU,Mem,Hdisk,Display_Card,MacAddress,'#','#',Import_time]
    return Infolist




if __name__ == '__main__':
    try:
        c = wmi.WMI()
        # 格式化pathlib返回的当前路径
        path_0 = str(pathlib.Path.cwd())
        print("运行该脚本将会在脚本目录下自动生成（电脑信息记录.xlsx）文件")
        print("如果已存在将会在里面继续写入信息")
        # 软件列表
        SoftList = SoftwareList()
        # 将hostname插入第0个索引 用于给sheet2添加 主机名
        SoftList.insert(0, printHostname())
        SoftList.insert(1, '')
        # #运行infolist来进行传参
        # Info_list()
        # excel名称 可以写入配置
        excel_name = '电脑信息记录.xlsx'
        # 创建的行 这个可以写入配置文件
        first_line = ['使用人', '计算机名', 'SN', '操作系统', 'CPU', '内存', '硬盘', '显卡', '网卡', '有线MAC', '无线MAC', '导入时间']
        while True:
            # 判断目录下是否存在该表格
            Excel_judge = os.path.isfile(path_0 + '\\' + excel_name)
            # 如果存在就运行excel_handle表格处理函数
            if Excel_judge == True:
                print('写入数据中...')

                excel_handle(Info_list(), SoftList, path_0, excel_name)
                break;
            else:
                print('文件不存在,生成中...')
                excel_new(first_line, excel_name)
                print('生成完毕！')

                continue;
    except Exception as e:
        print("Exception", e)
        input('运行主函数报错了！！！！')

