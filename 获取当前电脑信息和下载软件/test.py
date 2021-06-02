#!/usr/bin/env python 
# encoding: utf-8
import socket
import wmi
import pathlib
import time
from openpyxl import load_workbook

c = wmi.WMI()

#计算机名
def printHostname():
    hostname = socket.gethostname()
    # print(hostname)
    return hostname

#CPU
def printCPU():
    CPU = []
    for cpu in c.Win32_Processor():
        Cpu_dict = {}
        Cpu_dict['CPU型号'] = cpu.Name
        CPU.append(Cpu_dict)
    # print(CPU)
    return CPU

#BIOS
def printBIOS():
    BIOS = []
    for bios in c.Win32_BIOS():
        BIOS_dict = {}
        BIOS_dict['服务编码'] = bios.SerialNumber
        BIOS.append(BIOS_dict)
    # print(BIOS)
    return BIOS

#内存
def printPhysicalMemory():
    memorys = []

    for mem in c.Win32_PhysicalMemory():
        mem_dict = {}
        mem_dict['内存插槽'] = mem.DeviceLocator
        mem_dict['容量'] = int(int(mem.Capacity)/1024/1024)
        mem_dict['速度'] = mem.Speed
        mem_dict['厂商'] = mem.Manufacturer
        mem_dict['零件号'] = mem.PartNumber
        # MFRS = mem.Manufacturer
        # speed = mem.Speed
        # CY = mem.Capacity
        # PN = mem.PartNumber
        # DL = mem.DeviceLocator
        # print('设备位置: %s' %DL)
        # print('容量: %sMB' % (int(int(CY)/1024/1024)))
        # print('速度: %sMHz' % speed)
        # print('厂商: %s' % MFRS )
        # print('零件号: %s' % PN)
        memorys.append(mem_dict)
    # for m in memorys:
    #     print(m)
    # print(memorys)
    return memorys

#硬盘
def printDisk():
    disks = []
    for disk in c.Win32_DiskDrive():
        tmpmsg = {}
        tmpmsg['硬盘名称'] = disk.Caption
        tmpmsg['服务编码'] = disk.SerialNumber.strip()
        tmpmsg['容量'] = int(int(disk.Size)/1024/1024/1024)
        disks.append(tmpmsg)
    # for d in disks:
    #     print(d)
    # print(disks)
    return disks

#显卡
def printDisplay_Card():
    Display_Card = []
    for dp in c.Win32_VideoController():
        Display_Card_dict = {}
        Display_Card_dict['显卡名称'] = dp.Caption
        Display_Card.append(Display_Card_dict)
    # print(Display_Card)
    return Display_Card

#网卡mac地址
def printMacAddress():
    macs = []
    for n in  c.Win32_NetworkAdapter():
        mactmp = n.MACAddress
        if mactmp and len(mactmp.strip()) > 5:
            tmpmsg = {}
            tmpmsg['网卡名称'] = n.Name
            tmpmsg['MAC地址'] = n.MACAddress
            macs.append(tmpmsg)
    # print(macs)
    return macs

#操作系统
def printOS():
    OS = []
    for o in c.Win32_OperatingSystem():
        OS_dict = {}
        OS_dict['操作系统'] = o.Caption
        OS.append(OS_dict)
    # print(BIOS)
    return OS


#ecel表格处理
def excel_handle(list_all_def,path_0,excel_name):
    try:

        # 将信息写入excel表格
        wb = load_workbook(path_0 + '\\' + excel_name, keep_vba=True)

        # keep_vba=True 需要指定属性keep_vba=True 才能保存xlsm文档
        # wb = load_workbook(file1+'\\test_Excel.xlsm',keep_vba=True)
        # ws = wb.active                #获取当前活跃的sheet,默认是第一个sheetb

        sheets = wb.sheetnames  # 从名称获取sheet
        ws = wb[sheets[0]]
        # print(ws)

        ws.append(list_all_def)

        rows = ws.rows
        columns = ws.columns
        # 迭代所有的行
        # for row in rows:
        #     line = [col.value for col in row]
        #
        # # 通过坐标读取值
        #
        # cell_11 = ws.cell(row=1, column=1).value

        # 将此属性设置为False（默认），以保存为文档：
        wb.template = False

        # 您可以指定属性template = True，以将工作簿另存为模板：
        # wb.template = True

        wb.save(filename= path_0 + '\\' + excel_name)
        # wb.save(filename=file1+'\\test_Excel.xlsm')
        #
    except Exception as e:
        print("Exception", e)
        input('Please press enter key to exit ...')


if __name__ == '__main__':
    print("注意脚本运行路径下要存在（电脑信息记录.xlsm）文档")
    #使用人
    use_name = input("请输入用户姓名,以回车键结束。")
    # print(use_name)
    #计算机名
    hostname = printHostname()
    # print(hostname)
    #SN
    tmpSN = printBIOS()
    for i in tmpSN:
        SN = i['服务编码']
    # print(SN)
    #操作系统
    tmpOS = printOS()
    for i in tmpOS:
        OS = i['操作系统']
    # print(OS)
    #CPU
    tmpCPU = printCPU()
    for i in tmpCPU:
        CPU = i['CPU型号']
    # print(CPU)
    #内存
    # print(printPhysicalMemory())
    #硬盘
    # print(printDisk())
    #显卡
    # print(printDisplay_Card())
    #网卡
    # print(printMacAddress())
    #excel名称
    excel_name = '电脑信息记录.xlsm'
    #格式化pathlib返回的当前路径
    path_0 = str(pathlib.Path.cwd())
    #调用cominfo返回的 参数列表
    list_all_def =
    excel_handle(list_all_def,path_0,excel_name)
